"""Read part identifiers out of a customer BOM (Excel).

The sample BOM (bom_sample.xlsx) has columns:
    PartNumber/MPN/Comment | Package/Footprint | Quantity

For the crawler we only need the part identifier + quantity; header names vary
between customers, so we match columns loosely.
"""
from __future__ import annotations

from dataclasses import dataclass
from typing import List, Optional

from openpyxl import load_workbook

from .normalize import normalize_part_query

# Substrings we accept (case-insensitive) when locating each column.
_PART_HEADER_HINTS = ("part", "mpn", "p/n", "comment", "قطعه", "کالا")
_QTY_HEADER_HINTS = ("qty", "quantity", "تعداد")
_PKG_HEADER_HINTS = ("package", "footprint", "بسته")

# Worksheet titles that are not BOMs (index / metadata / TOC sheets) — skipped
# when reading a whole multi-sheet workbook.
_SHEET_SKIP_HINTS = ("index", "metadata", "meta", "راهنما", "فهرست", "خلاصه")


@dataclass
class BomLine:
    part: str            # normalized query string
    raw_part: str        # exactly as written in the sheet
    quantity: Optional[float] = None
    package: Optional[str] = None
    row: int = 0
    sheet: str = ""      # worksheet the line came from (for multi-BOM workbooks)


def _find_col(headers: List[str], hints) -> Optional[int]:
    for idx, h in enumerate(headers):
        low = (h or "").strip().lower()
        if any(hint in low for hint in hints):
            return idx
    return None


def _read_worksheet(ws, sheet_name: str = "") -> List[BomLine]:
    """Parse one worksheet into ``BomLine`` records (blank part rows skipped)."""
    rows = ws.iter_rows(values_only=True)
    try:
        header = [str(c) if c is not None else "" for c in next(rows)]
    except StopIteration:
        return []

    part_col = _find_col(header, _PART_HEADER_HINTS)
    if part_col is None:
        part_col = 0  # fall back to the first column
    qty_col = _find_col(header, _QTY_HEADER_HINTS)
    pkg_col = _find_col(header, _PKG_HEADER_HINTS)

    lines: List[BomLine] = []
    for i, row in enumerate(rows, start=2):  # row 1 was the header
        raw = row[part_col] if part_col < len(row) else None
        if raw is None or str(raw).strip() == "":
            continue
        raw_part = str(raw).strip()
        quantity = None
        if qty_col is not None and qty_col < len(row) and row[qty_col] is not None:
            try:
                quantity = float(row[qty_col])
            except (TypeError, ValueError):
                quantity = None
        package = None
        if pkg_col is not None and pkg_col < len(row) and row[pkg_col] is not None:
            package = str(row[pkg_col]).strip()

        lines.append(
            BomLine(
                part=normalize_part_query(raw_part),
                raw_part=raw_part,
                quantity=quantity,
                package=package,
                row=i,
                sheet=sheet_name,
            )
        )
    return lines


def read_bom(path: str, sheet: Optional[str] = None) -> List[BomLine]:
    """Parse a single BOM sheet into ``BomLine`` records."""
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[sheet] if sheet else wb.worksheets[0]
        return _read_worksheet(ws, ws.title)
    finally:
        wb.close()


def read_all_sheets(
    path: str,
    sheets: Optional[List[str]] = None,
    per_sheet: Optional[int] = None,
) -> List[BomLine]:
    """Parse every BOM sheet of a multi-sheet workbook.

    ``sheets`` selects specific worksheet titles; otherwise all sheets are used
    except index/metadata sheets (matched by :data:`_SHEET_SKIP_HINTS`).
    ``per_sheet`` caps how many part-lines are taken from each sheet (useful to
    sample every board without crawling thousands of parts). Each line records
    its source ``sheet``.
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        if sheets:
            wanted = [str(s) for s in sheets]
            names = [n for n in wb.sheetnames if n in wanted]
        else:
            names = [
                n for n in wb.sheetnames
                if not any(h in n.lower() for h in _SHEET_SKIP_HINTS)
            ]
        out: List[BomLine] = []
        for name in names:
            lines = _read_worksheet(wb[name], name)
            if per_sheet:
                lines = lines[:per_sheet]
            out.extend(lines)
        return out
    finally:
        wb.close()


def unique_parts(lines: List[BomLine]) -> List[str]:
    """Distinct, order-preserving list of part queries to crawl."""
    seen: set[str] = set()
    out: List[str] = []
    for line in lines:
        if line.part and line.part not in seen:
            seen.add(line.part)
            out.append(line.part)
    return out
