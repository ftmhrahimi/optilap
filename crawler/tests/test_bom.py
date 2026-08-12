"""Tests for BOM reading against the real sample workbook when available."""
from pathlib import Path

import pytest

from optilap_crawler.bom import read_bom, unique_parts

# The sample BOM lives outside the repo (attached to the task). Tests that need
# it are skipped when it isn't present so CI stays green.
SAMPLE = Path(__file__).resolve().parent.parent / "fixtures" / "bom_sample.xlsx"


@pytest.mark.skipif(not SAMPLE.exists(), reason="bom_sample.xlsx not present")
def test_read_sample_bom():
    lines = read_bom(str(SAMPLE))
    assert lines, "expected at least one BOM line"
    parts = {l.raw_part for l in lines}
    assert "LM358DR" in parts or "ATMEGA16A-AU" in parts
    # Quantities parse as numbers.
    assert any(l.quantity is not None for l in lines)


@pytest.mark.skipif(not SAMPLE.exists(), reason="bom_sample.xlsx not present")
def test_unique_parts_dedupes():
    lines = read_bom(str(SAMPLE))
    uniq = unique_parts(lines)
    assert len(uniq) == len(set(uniq))
    assert len(uniq) <= len(lines)


def test_read_all_sheets_multisheet(tmp_path):
    """read_all_sheets pulls parts from every BOM sheet, skips metadata sheets,
    tags each line with its source sheet, and honours per_sheet sampling."""
    from openpyxl import Workbook
    from optilap_crawler.bom import read_all_sheets

    wb = Workbook()
    meta = wb.active
    meta.title = "Index_Metadata"
    meta.append(["Sheet Code", "Board Name"])
    meta.append(["BOM_01", "Power board"])

    b1 = wb.create_sheet("1")
    b1.append(["ردیف", "نام قطعه / P/N", "تعداد (Qty)"])
    b1.append([1, "STM32F103C8T6", 1])
    b1.append([2, "LM7805", 2])

    b2 = wb.create_sheet("2")
    b2.append(["#", "MPN", "Qty"])
    b2.append([1, "NE555", 3])
    b2.append([2, "LM7805", 1])   # duplicate across sheets
    b2.append([3, "IRF3205", 1])

    path = tmp_path / "multi.xlsx"
    wb.save(path)

    lines = read_all_sheets(str(path))
    raws = {l.raw_part for l in lines}
    assert "STM32F103C8T6" in raws and "NE555" in raws and "IRF3205" in raws
    # Metadata sheet skipped (its "Board Name" values must not leak in as parts).
    assert "Power board" not in raws
    # Each line carries its source sheet.
    assert {l.sheet for l in lines} == {"1", "2"}
    # Dedup across sheets collapses the shared LM7805.
    assert sorted(unique_parts(lines)) == sorted({"STM32F103C8T6", "LM7805",
                                                  "NE555", "IRF3205"})

    # per_sheet caps lines taken from each sheet.
    capped = read_all_sheets(str(path), per_sheet=1)
    assert len([l for l in capped if l.sheet == "1"]) == 1
    assert len([l for l in capped if l.sheet == "2"]) == 1

    # Explicit sheet selection.
    only2 = read_all_sheets(str(path), sheets=["2"])
    assert {l.sheet for l in only2} == {"2"}
